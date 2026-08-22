from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
from itertools import combinations_with_replacement
from pathlib import Path
import re
import unicodedata

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def _text(value: object) -> str:
    if value is None:
        return ""
    return str(value).replace("\xa0", " ").replace("\\_", "_").strip()


def _key(value: object) -> str:
    return unicodedata.normalize("NFKD", _text(value)).encode("ascii", "ignore").decode("ascii").casefold()


def _has(text: str, *terms: str) -> bool:
    return any(re.search(r"(?<!\w)" + re.escape(_key(term)) + r"(?!\w)", text) for term in terms)


ACTION_TERMS: tuple[tuple[str, tuple[str, ...]], ...] = (
    ("Administração de acessos", ("criar usuario", "alterar usuario", "criar perfil", "alterar perfil", "conceder acesso", "remover acesso", "administrar grupo", "administrar usuario", "administrar perfil", "recertificar acesso", "revisar acesso", "create user", "update user", "grant access", "remove access", "manage group", "review access")),
    ("Auditoria/controle", ("auditar", "auditoria", "investigar", "certificar", "monitorar", "audit", "investigate", "certify", "monitor")),
    ("Parametrização/configuração", ("parametrizar", "configurar", "definir regra", "alterar limite", "alterar alcada", "alterar tarifa", "duplo comando", "parameterize", "configure", "set rule", "change limit", "change threshold")),
    ("Desbloqueio", ("desbloquear", "remover restricao", "liberar posicao", "liberar saldo", "unblock", "remove restriction")),
    ("Bloqueio", ("bloquear", "aplicar restricao", "block", "apply restriction")),
    ("Estorno", ("estornar", "reverter", "reverse", "reversal")),
    ("Cancelamento", ("cancelar", "reativar", "cancel", "reactivate")),
    ("Exclusão", ("excluir", "deletar", "delete", "remove")),
    ("Assinatura", ("assinar", "assinatura", "sign", "signature")),
    ("Autorização", ("autorizar", "autorizacao", "authorize", "authorise", "authorization")),
    ("Aprovação", ("aprovar", "aprovacao", "homologar", "approve", "approval")),
    ("Liberação", ("liberar", "liberacao", "release")),
    ("Validação", ("validar", "validacao", "conferir", "conferencia", "atestar", "validate", "validation", "verify", "check", "attest")),
    ("Conciliação", ("conciliar", "conciliacao", "encerrar conciliacao", "reconcile", "reconciliation")),
    ("Execução/processamento", ("executar pagamento", "processar lote", "efetuar transferencia", "movimentar", "liquidar", "contabilizar", "efetivar", "fechar operacao", "realizar lancamento", "execute", "process", "transfer", "pay", "settle", "post transaction")),
    ("Inclusão/cadastro", ("cadastrar", "incluir", "criar", "registrar", "emitir", "inserir", "add", "create", "register", "insert", "issue")),
    ("Alteração/manutenção", ("alterar", "manter", "corrigir", "ajustar", "gerenciar dados", "change", "update", "maintain", "correct", "adjust", "manage data")),
    ("Solicitação", ("solicitar", "iniciar operacao", "registrar pedido", "criar requisicao", "criar pedido", "request", "submit", "initiate operation")),
    ("Consulta", ("consultar", "consulta", "visualizar", "relatorio", "view", "query", "report")),
)


OBJECTS: tuple[tuple[str, str, tuple[str, ...]], ...] = (
    ("Gestão de acessos", "Acesso", ("usuario", "usuarios", "perfil", "perfis", "grupo", "grupos", "acesso", "acessos")),
    ("Gestão de logs", "Log", ("log", "logs", "trilha", "evidencia", "evidencias", "retencao")),
    ("Dados mestres", "Dados mestres", ("cliente", "fornecedor", "beneficiario", "conta bancaria", "participante", "carteira", "dado mestre")),
    ("Operações financeiras", "Movimentação financeira", ("pagamento", "transferencia", "movimentar ativo", "movimentar saldo", "resgatar", "liquidar", "saldo")),
    ("Aquisições", "Aquisição", ("requisicao", "pedido de compra", "recebimento", "nota fiscal")),
    ("Documentos", "Documento", ("documento", "contrato", "titulo", "ordem", "guia")),
    ("Conciliação", "Conciliação", ("conciliacao", "extrato", "movimento", "inconsistencia")),
    ("Crédito", "Crédito", ("credito", "limite", "proposta")),
    ("Parametrização", "Parâmetro", ("parametro", "tarifa", "alcada", "regra de aprovacao")),
    ("Controles preventivos", "Restrição", ("bloqueio", "bloquear", "desbloquear", "restricao")),
)


@dataclass(frozen=True)
class RuleDefinition:
    rule_id: str
    name: str
    left_actions: tuple[str, ...]
    right_actions: tuple[str, ...]
    risk: str
    severity: str
    objects: tuple[tuple[str, str], ...] = ()


RULES: tuple[RuleDefinition, ...] = (
    RuleDefinition("SOD_001", "Solicitar × aprovar", ("Solicitação", "Inclusão/cadastro"), ("Aprovação", "Autorização", "Assinatura", "Liberação"), "O usuário pode iniciar e decidir sobre a própria solicitação.", "Crítico"),
    RuleDefinition("SOD_002", "Manter × validar", ("Inclusão/cadastro", "Alteração/manutenção", "Exclusão"), ("Aprovação", "Autorização", "Validação", "Liberação"), "O usuário pode alterar os dados que ele próprio aprova.", "Crítico"),
    RuleDefinition("SOD_003", "Executar × validar", ("Execução/processamento",), ("Validação", "Aprovação"), "O usuário pode executar e validar o próprio processamento.", "Crítico"),
    RuleDefinition("SOD_004", "Parametrizar × operar", ("Parametrização/configuração",), ("Execução/processamento", "Aprovação", "Autorização", "Liberação"), "O usuário pode modificar o controle e executar a operação afetada.", "Crítico"),
    RuleDefinition("SOD_005", "Dados mestres × pagamento ou transferência", ("Inclusão/cadastro", "Alteração/manutenção"), ("Execução/processamento",), "O usuário pode alterar o destino ou beneficiário e realizar a movimentação.", "Crítico", (("Dados mestres", "Movimentação financeira"), ("Movimentação financeira", "Dados mestres"))),
    RuleDefinition("SOD_006", "Preparar × assinar", ("Inclusão/cadastro", "Alteração/manutenção"), ("Assinatura", "Aprovação", "Autorização", "Liberação"), "O usuário pode preparar e formalizar o próprio documento.", "Alto", (("Documento", "Documento"),)),
    RuleDefinition("SOD_007", "Operação financeira × conciliação", ("Execução/processamento",), ("Conciliação", "Validação"), "O usuário pode executar e conciliar a própria transação.", "Crítico", (("Movimentação financeira", "Conciliação"), ("Conciliação", "Movimentação financeira"))),
    RuleDefinition("SOD_008", "Bloquear × desbloquear", ("Bloqueio",), ("Desbloqueio", "Liberação"), "O usuário pode aplicar e remover o próprio controle preventivo.", "Alto", (("Restrição", "Restrição"),)),
    RuleDefinition("SOD_009", "Conceder acesso × aprovar ou revisar acesso", ("Administração de acessos",), ("Aprovação", "Autorização", "Auditoria/controle"), "O usuário pode conceder, aprovar ou revisar acessos sem independência.", "Crítico", (("Acesso", "Acesso"),)),
    RuleDefinition("SOD_010", "Administrar logs × auditar", ("Alteração/manutenção", "Administração de acessos"), ("Auditoria/controle", "Validação"), "O usuário pode alterar a evidência que deveria auditar.", "Crítico", (("Log", "Log"),)),
    RuleDefinition("SOD_011", "Dados de crédito × aprovar crédito", ("Inclusão/cadastro", "Alteração/manutenção"), ("Aprovação", "Autorização", "Liberação"), "O usuário pode manipular os dados usados na própria decisão de crédito.", "Crítico", (("Crédito", "Crédito"),)),
    RuleDefinition("SOD_012", "Criar × cancelar ou estornar", ("Inclusão/cadastro", "Execução/processamento"), ("Cancelamento", "Estorno", "Exclusão"), "O cancelamento ou estorno pode ocultar a própria operação.", "Alto", (("Movimentação financeira", "Movimentação financeira"), ("Aquisição", "Aquisição"), ("Documento", "Documento"), ("Crédito", "Crédito"))),
)


@dataclass
class SodAnalysisResult:
    template_conflicts: list[dict]
    conflicts: list[dict]
    classifications: list[dict]
    rules: list[dict]
    embedded: list[dict]
    gaps: list[dict]
    user_impacted: list[dict]
    no_direct: list[dict]
    sat_candidates: list[dict]
    summary: dict[str, int]


def _classify_action(text: str) -> str:
    vague = ("executar funcionalidade", "acessar tela", "manutencao", "administracao", "ponto de verificacao", "processamento", "controle")
    for name, terms in ACTION_TERMS:
        if _has(text, *terms):
            return name
    return "Indeterminada" if _has(text, *vague) or text else "Indeterminada"


def _process_object(text: str) -> tuple[str, str]:
    for process, obj, terms in OBJECTS:
        if _has(text, *terms):
            return process, obj
    return "Não identificado", "Não identificado"


def _combined_text(row: dict) -> str:
    return _key(" ".join(_text(row.get(field)) for field in ("atividade", "funcionalidade", "transacao_tela_direito", "modulo")))


def _append_distinct(current: object, *values: object) -> str:
    return " | ".join(sorted({_text(value) for value in (current, *values) if _text(value)}))


SEMANTIC_STOPWORDS = {
    "a", "as", "ao", "aos", "com", "da", "das", "de", "do", "dos", "e", "em", "na", "nas", "no", "nos", "o", "os", "ou", "para", "por",
    "the", "and", "for", "of", "to", "in", "on", "with", "from",
    "atividade", "atividades", "funcao", "funcionalidade", "funcionalidades", "sistema", "sistemas", "transacao", "transacoes", "tela", "telas", "direito", "direitos",
    "operacao", "operacoes", "processo", "processos", "informacao", "informacoes", "geral", "gestao", "gestionar", "cac",
}
SEMANTIC_ACTION_WORDS = {
    word for _, terms in ACTION_TERMS for term in terms for word in _key(term).split() if len(word) >= 3
}


def _semantic_tokens(text: str) -> set[str]:
    """Retorna termos de negócio que sustentam a comparação entre as duas ações.

    A inferência só é usada após os verbos de uma regra SoD terem sido identificados.
    Assim, ela não cria conflito apenas por duas atividades terem o mesmo perfil ou sistema.
    """
    return {
        token for token in re.findall(r"[a-z0-9]{3,}", text)
        if token not in SEMANTIC_STOPWORDS and token not in SEMANTIC_ACTION_WORDS
    }


def _compatible(rule: RuleDefinition, left: dict, right: dict) -> bool:
    left_object, right_object = left["Objeto de negócio"], right["Objeto de negócio"]
    if rule.objects:
        if (left_object, right_object) in rule.objects:
            return True
    elif left["Processo"] == right["Processo"] and left_object == right_object and left_object != "Não identificado":
        return True
    # Quando a taxonomia ainda não conhece o objeto, exige um termo de negócio
    # comum e explícito nas duas descrições. Isso torna a regra aplicável a novos
    # sistemas sem transformar o nome de um sistema (por exemplo, CAC) em regra.
    return bool(left.get("_termos_negocio", set()) & right.get("_termos_negocio", set()))


def _style(sheet) -> None:
    header = PatternFill("solid", fgColor="1F1F1F")
    title = PatternFill("solid", fgColor="3A3A3A")
    accent = PatternFill("solid", fgColor="F4B183")
    sheet.freeze_panes = "A3"
    sheet.sheet_view.showGridLines = False
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, sheet.max_column))
    cell = sheet.cell(1, 1)
    cell.fill = title
    cell.font = Font(color="FFFFFF", bold=True, size=14)
    cell.alignment = Alignment(horizontal="left")
    for cell in sheet[2]:
        cell.fill = header
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    for row in sheet.iter_rows(min_row=3):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for index in range(1, sheet.max_column + 1):
        values = [_text(sheet.cell(row, index).value) for row in range(1, min(sheet.max_row, 100) + 1)]
        sheet.column_dimensions[get_column_letter(index)].width = min(42, max(12, max((len(value) for value in values), default=12) + 2))
    if sheet.max_row >= 2:
        sheet.row_dimensions[2].height = 32
    if sheet.max_row >= 3 and sheet.title in {"Matriz de Conflitos", "Lacunas e Validações"}:
        for cell in sheet[3]:
            cell.fill = accent


def _write_sheet(book: Workbook, name: str, title: str, columns: list[str], rows: list[dict]) -> None:
    sheet = book.create_sheet(name)
    sheet.append([title])
    sheet.append(columns)
    for row in rows:
        sheet.append([row.get(column, "") for column in columns])
    _style(sheet)


def export_analysis(path: Path, result: SodAnalysisResult) -> None:
    book = Workbook()
    book.remove(book.active)
    summary_rows = [{"Indicador": key.replace("_", " ").title(), "Quantidade": value} for key, value in result.summary.items()]
    _write_sheet(book, "Resumo Executivo", "Resumo Executivo — análise conservadora SoD", ["Indicador", "Quantidade"], summary_rows)
    _write_sheet(book, "Matriz Perfil x Perfil", "Matriz Perfil x Perfil", ["Sistema", "Perfil A", "Perfil B", "Regra SoD", "Tipo", "Severidade", "Status"], result.conflicts)
    _write_sheet(book, "Matriz de Conflitos", "Matriz de Conflitos", list(result.conflicts[0].keys()) if result.conflicts else ["ID", "Sistema", "Regra SoD", "Status"], result.conflicts)
    _write_sheet(book, "Regras de Conflito", "Regras de Conflito", ["ID da regra", "Sistema", "Regra SoD", "Processo", "Objeto de negócio", "Direitos do Lado A", "Direitos do Lado B", "Risco", "Severidade", "Evidência", "Status", "Tratamento recomendado", "Validação necessária"], result.rules)
    _write_sheet(book, "Classificação Funcional", "Classificação Funcional", list(result.classifications[0].keys()) if result.classifications else ["ID", "Sistema", "Perfil", "Classificação da ação"], result.classifications)
    _write_sheet(book, "Conflitos por Perfil", "Conflitos por Perfil", ["Sistema", "Perfil A", "Perfil B", "Regra SoD", "Severidade", "Status"], result.conflicts)
    _write_sheet(book, "Matriz Funcional Fonte", "Matriz Funcional Fonte", ["ID", "Sistema", "Perfil", "Atividade", "Funcionalidade", "Transação/tela/direito", "Linha da fonte", "Arquivo/aba de origem"], result.classifications)
    _write_sheet(book, "Sem Conflito Direto", "Sem Conflito Direto", ["Sistema", "Perfil A", "Perfil B", "Resultado"], result.no_direct)
    _write_sheet(book, "Conflitos Embutidos", "Conflitos Embutidos", ["Sistema", "Perfil", "Direito", "Conflito embutido/intrapermissão", "Linha da fonte", "Arquivo/aba de origem", "Status", "Tratamento recomendado"], result.embedded)
    _write_sheet(book, "Usuários Impactados", "Usuários Impactados", ["Usuário", "Sistema", "Perfis", "Regra SoD", "Tipo", "Status"], result.user_impacted)
    _write_sheet(book, "Lacunas e Validações", "Lacunas e Validações", ["Tipo", "Sistema", "Perfil", "Direito", "Linha da fonte", "Arquivo/aba de origem", "Pendência", "Tratamento"], result.gaps)
    _write_sheet(book, "Critérios SAT", "Critérios SAT — derivados da Matriz Funcional e SoD", ["ID", "Sistema", "Perfil", "Atividade", "Classificação SAT sugerida", "Justificativa SAT", "Risco associado", "Validação adicional"], result.sat_candidates)
    criteria = [
        {"Critério": "Fonte metodológica", "Aplicação": "Raciocínio SoD fornecido pelo usuário nesta solicitação."},
        {"Critério": "Status padrão", "Aplicação": "A validar; Confirmado exige workflow e concentração pela mesma identidade documentados."},
        {"Critério": "Ação indeterminada", "Aplicação": "Não produz conflito confirmado; é registrada como lacuna."},
        {"Critério": "Deduplicação", "Aplicação": "Somente registros absolutamente idênticos podem ser removidos na fonte; pares invertidos são consolidados."},
        {"Critério": "SAT crítica", "Aplicação": "Atividade participante de conflito SoD potencial ou ação explícita de parametrização, administração de acessos, aprovação, autorização, assinatura, liberação ou execução/processamento."},
        {"Critério": "SAT alta", "Aplicação": "Ação explícita de exclusão, cancelamento, estorno, bloqueio, desbloqueio, validação, conciliação ou auditoria/controle."},
    ]
    _write_sheet(book, "Critérios e Qualidade", "Critérios e Qualidade", ["Critério", "Aplicação"], criteria)
    book.save(path)


def analyze(matrix: list[dict], contexts: list[dict]) -> SodAnalysisResult:
    classifications: list[dict] = []
    gaps: list[dict] = []
    embedded: list[dict] = []
    for row in matrix:
        text = _combined_text(row)
        action = _classify_action(text)
        process, obj = _process_object(text)
        source = f"{_text(row.get('_arquivo_origem'))} / aba conforme log da execução"
        item = {"ID": _text(row.get("id")), "Sistema": _text(row.get("sistema")), "Perfil": _text(row.get("perfil")), "Atividade": _text(row.get("atividade")), "Funcionalidade": _text(row.get("funcionalidade")), "Transação/tela/direito": _text(row.get("transacao_tela_direito")), "Processo": process, "Objeto de negócio": obj, "Classificação da ação": action, "Natureza do acesso": "Não inferida", "Potencial de alteração": "Não inferido", "Potencial de aprovação": "Não inferido", "Potencial de efeito": "Não inferido", "Evidência": "Verbo/objeto identificados literalmente na fonte" if action != "Indeterminada" and obj != "Não identificado" else "Informação insuficiente", "Linha da fonte": row.get("_linha_origem", ""), "Arquivo/aba de origem": source}
        classifications.append(item)
        if action == "Indeterminada" or obj == "Não identificado":
            gaps.append({"Tipo": "Classificação insuficiente", "Sistema": item["Sistema"], "Perfil": item["Perfil"], "Direito": item["Transação/tela/direito"], "Linha da fonte": item["Linha da fonte"], "Arquivo/aba de origem": source, "Pendência": "Ação ou objeto não está explícito o suficiente para inferência conservadora.", "Tratamento": "Validar descrição funcional ou workflow; não tratar como conflito confirmado."})
        for left, right, label in (("bloquear", "desbloquear", "Bloquear × desbloquear"), ("solicitar", "aprovar", "Solicitar × aprovar"), ("executar", "validar", "Executar × validar"), ("conceder", "revisar", "Conceder acesso × revisar")):
            if _has(text, left) and _has(text, right):
                embedded.append({"Sistema": item["Sistema"], "Perfil": item["Perfil"], "Direito": item["Transação/tela/direito"], "Conflito embutido/intrapermissão": label, "Linha da fonte": item["Linha da fonte"], "Arquivo/aba de origem": source, "Status": "A validar", "Tratamento recomendado": "Decompor tecnicamente a permissão; não foi criado par artificial do direito com ele mesmo."})

    indexed = [dict(item, _sistema=_key(item["Sistema"]), _perfil=_text(item["Perfil"]), _termos_negocio=_semantic_tokens(_key(" ".join((item["Atividade"], item["Funcionalidade"], item["Transação/tela/direito"])))) ) for item in classifications if item["Sistema"] and item["Perfil"]]
    conflicts_by_key: dict[tuple[str, str, str, str], dict] = {}
    rules_by_key: dict[tuple[str, str, str, str], dict] = {}
    template_conflicts: list[dict] = []
    conflict_activity_ids: set[str] = set()
    for rule in RULES:
        lefts = [row for row in indexed if row["Classificação da ação"] in rule.left_actions]
        rights = [row for row in indexed if row["Classificação da ação"] in rule.right_actions]
        for left in lefts:
            for right in rights:
                if left["_sistema"] != right["_sistema"] or left["ID"] == right["ID"] or not _compatible(rule, left, right):
                    continue
                conflict_activity_ids.update((left["ID"], right["ID"]))
                profile_a, profile_b = sorted((left["Perfil"], right["Perfil"]))
                key = (left["_sistema"], rule.rule_id, profile_a, profile_b)
                conflict = conflicts_by_key.setdefault(key, {"ID": f"{rule.rule_id}_{len(conflicts_by_key) + 1:05d}", "Sistema": left["Sistema"], "Regra SoD": rule.name, "Processo": left["Processo"], "Objeto de negócio": f"{left['Objeto de negócio']} × {right['Objeto de negócio']}" if left["Objeto de negócio"] != right["Objeto de negócio"] else left["Objeto de negócio"], "Perfil A": profile_a, "Papel SoD do Perfil A": "Lado A" if profile_a == left["Perfil"] else "Lado B", "Direito(s) do Perfil A": "", "Perfil B": profile_b, "Papel SoD do Perfil B": "Lado B" if profile_b == right["Perfil"] else "Lado A", "Direito(s) do Perfil B": "", "ID Atividade(s) Lado A": "", "Atividade(s) Lado A": "", "ID Atividade(s) Lado B": "", "Atividade(s) Lado B": "", "Tipo": "Intraperfil" if profile_a == profile_b else "Entre perfis", "Severidade": rule.severity, "Evidência": "Alta — verbos e objeto compatível estão explícitos na fonte; a concentração pela mesma identidade requer validação." if left["Objeto de negócio"] != "Não identificado" else "Média — ações complementares com termo de negócio comum; confirmar workflow, escopo e alçada.", "Status": "A validar", "Descrição do risco": rule.risk, "Tratamento recomendado": "Não revogar automaticamente; validar workflow, escopo e concentração de perfis pela mesma identidade.", "Evidência necessária para validação": "Workflow, escopo do direito e vínculo usuário × perfil vigente.", "Linha(s) da fonte": "", "Arquivo/aba de origem": ""})
                side_a, side_b = left, right
                if conflict["Perfil A"] == side_a["Perfil"]:
                    conflict["Direito(s) do Perfil A"] = _append_distinct(conflict["Direito(s) do Perfil A"], side_a["Transação/tela/direito"])
                    conflict["Direito(s) do Perfil B"] = _append_distinct(conflict["Direito(s) do Perfil B"], side_b["Transação/tela/direito"])
                else:
                    conflict["Direito(s) do Perfil A"] = _append_distinct(conflict["Direito(s) do Perfil A"], side_b["Transação/tela/direito"])
                    conflict["Direito(s) do Perfil B"] = _append_distinct(conflict["Direito(s) do Perfil B"], side_a["Transação/tela/direito"])
                conflict["ID Atividade(s) Lado A"] = _append_distinct(conflict["ID Atividade(s) Lado A"], side_a["ID"])
                conflict["Atividade(s) Lado A"] = _append_distinct(conflict["Atividade(s) Lado A"], side_a["Atividade"])
                conflict["ID Atividade(s) Lado B"] = _append_distinct(conflict["ID Atividade(s) Lado B"], side_b["ID"])
                conflict["Atividade(s) Lado B"] = _append_distinct(conflict["Atividade(s) Lado B"], side_b["Atividade"])
                conflict["Linha(s) da fonte"] = _append_distinct(conflict["Linha(s) da fonte"], left["Linha da fonte"], right["Linha da fonte"])
                conflict["Arquivo/aba de origem"] = _append_distinct(conflict["Arquivo/aba de origem"], left["Arquivo/aba de origem"], right["Arquivo/aba de origem"])
                rule_key = (left["_sistema"], rule.rule_id, left["Processo"], conflict["Objeto de negócio"])
                specific = rules_by_key.setdefault(rule_key, {"ID da regra": rule.rule_id, "Sistema": left["Sistema"], "Regra SoD": rule.name, "Processo": left["Processo"], "Objeto de negócio": conflict["Objeto de negócio"], "Direitos do Lado A": "", "Direitos do Lado B": "", "Risco": rule.risk, "Severidade": rule.severity, "Evidência": "Atividades aderentes identificadas literalmente na matriz funcional.", "Status": "A validar", "Tratamento recomendado": "Validar workflow e segregação da atribuição antes de qualquer tratamento de acesso.", "Validação necessária": "Confirmação funcional do objeto, escopo e alçadas."})
                specific["Direitos do Lado A"] = _append_distinct(specific["Direitos do Lado A"], side_a["Transação/tela/direito"])
                specific["Direitos do Lado B"] = _append_distinct(specific["Direitos do Lado B"], side_b["Transação/tela/direito"])

    conflicts = sorted(conflicts_by_key.values(), key=lambda row: (row["Sistema"], row["Perfil A"], row["Perfil B"], row["Regra SoD"]))
    for conflict in conflicts:
        template_conflicts.append({"id_conflito": conflict["ID"], "id_atividade_1": conflict["ID Atividade(s) Lado A"], "atividade_1": conflict["Atividade(s) Lado A"], "id_atividade_2": conflict["ID Atividade(s) Lado B"], "atividade_2": conflict["Atividade(s) Lado B"], "risco_descricao": conflict["Descrição do risco"], "justificativa_conflito": conflict["Evidência"], "classificacao_risco": conflict["Severidade"], "recomendacao": conflict["Tratamento recomendado"], "conflito": "Potencial", "identificacao": conflict["Status"], "legenda": "Metodologia SoD fornecida pelo usuário"})

    user_profiles: dict[str, dict[str, set[str]]] = defaultdict(lambda: defaultdict(set))
    for context in contexts:
        system = context["sistema_funcionalidades"]
        for row in context["usuarios"]:
            user, profile = _text(row.get("usuario")), _text(row.get("perfil"))
            if user and profile:
                user_profiles[system][user].add(profile)
    user_impacted: list[dict] = []
    for conflict in conflicts:
        system = _key(conflict["Sistema"])
        for user, profiles in user_profiles.get(system, {}).items():
            if conflict["Perfil A"] in profiles and conflict["Perfil B"] in profiles:
                user_impacted.append({"Usuário": user, "Sistema": conflict["Sistema"], "Perfis": " | ".join(sorted(profiles)), "Regra SoD": conflict["Regra SoD"], "Tipo": conflict["Tipo"], "Status": "A validar"})

    profiles_by_system: dict[str, set[str]] = defaultdict(set)
    display_system: dict[str, str] = {}
    for row in indexed:
        profiles_by_system[row["_sistema"]].add(row["Perfil"])
        display_system.setdefault(row["_sistema"], row["Sistema"])
    direct_pairs = {(key[0], key[2], key[3]) for key in conflicts_by_key}
    no_direct: list[dict] = []
    maximum_pairs_exported = 200_000
    for system, profiles in profiles_by_system.items():
        for profile_a, profile_b in combinations_with_replacement(sorted(profiles), 2):
            if (system, profile_a, profile_b) not in direct_pairs:
                if len(no_direct) < maximum_pairs_exported:
                    no_direct.append({"Sistema": display_system[system], "Perfil A": profile_a, "Perfil B": profile_b, "Resultado": "Sem conflito direto nas regras SoD identificadas; não representa aprovação automática do acesso."})
                elif not any(gap["Tipo"] == "Limite operacional de exportação" for gap in gaps):
                    gaps.append({"Tipo": "Limite operacional de exportação", "Sistema": display_system[system], "Perfil": "", "Direito": "", "Linha da fonte": "", "Arquivo/aba de origem": "", "Pendência": "A aba Sem Conflito Direto ultrapassou 200.000 pares e foi limitada para preservar a execução local.", "Tratamento": "Use o resumo e processe o sistema isoladamente para exportação integral."})
    unique_pairs = sum(len(profiles) * (len(profiles) + 1) // 2 for profiles in profiles_by_system.values())
    critical_actions = {"Parametrização/configuração", "Administração de acessos", "Aprovação", "Autorização", "Assinatura", "Liberação", "Execução/processamento"}
    high_actions = {"Exclusão", "Cancelamento", "Estorno", "Bloqueio", "Desbloqueio", "Validação", "Conciliação", "Auditoria/controle"}
    sat_candidates: list[dict] = []
    for row in classifications:
        action, identifier = row["Classificação da ação"], row["ID"]
        if identifier in conflict_activity_ids:
            level, reason, risk = "Crítica", "Participa de conflito funcional potencial identificado pelas regras SoD metodológicas.", "Concentração de etapas incompatíveis do processo."
        elif action in critical_actions:
            level, reason, risk = "Crítica", f"Ação explícita classificada como {action} na Matriz Funcional.", "Ação crítica que exige restrição e validação de alçada."
        elif action in high_actions:
            level, reason, risk = "Alta", f"Ação explícita classificada como {action} na Matriz Funcional.", "Ação sensível que requer controle independente e revisão periódica."
        else:
            continue
        sat_candidates.append({"id": f"SAT_{identifier}", "area": "", "atividade": row["Atividade"], "descricao_atividade": row["Atividade"], "perfil": row["Perfil"], "funcionalidade": row["Funcionalidade"], "modulo": "", "transacao_tela_direito": row["Transação/tela/direito"], "classificacao_sat_sugerida": f"{level} — A validar", "justificativa_sat": reason, "risco_associado": risk, "restricao_acesso": "Conceder somente a perfis autorizados e com necessidade de negócio comprovada.", "controle_mitigatorio": "Aprovação independente, registro de trilha e revisão periódica do acesso.", "controle_compensatorio": "Monitoramento posterior e evidência de revisão independente, quando a segregação não for viável.", "validacao_adicional": f"Validar workflow, alçada, escopo organizacional e linha da fonte {row['Linha da fonte']}.", "sistema": row["Sistema"], "ID": identifier, "Sistema": row["Sistema"], "Perfil": row["Perfil"], "Atividade": row["Atividade"], "Classificação SAT sugerida": f"{level} — A validar", "Justificativa SAT": reason, "Risco associado": risk, "Validação adicional": f"Validar workflow, alçada, escopo organizacional e linha da fonte {row['Linha da fonte']}."})
    summary = {"sistemas": len(profiles_by_system), "perfis": sum(len(values) for values in profiles_by_system.values()), "direitos": len(classifications), "regras_sod": len(rules_by_key), "ocorrencias": len(conflicts), "pares_unicos": unique_pairs, "intraperfil": sum(row["Tipo"] == "Intraperfil" for row in conflicts), "entre_perfis": sum(row["Tipo"] == "Entre perfis" for row in conflicts), "critico": sum(row["Severidade"] == "Crítico" for row in conflicts), "alto": sum(row["Severidade"] == "Alto" for row in conflicts), "medio": sum(row["Severidade"] == "Médio" for row in conflicts), "sem_conflito_direto": len(no_direct), "direitos_indeterminados": sum(row["Classificação da ação"] == "Indeterminada" for row in classifications), "atividades_sat": len(sat_candidates), "usuarios_impactados": len(user_impacted), "lacunas": len(gaps)}
    return SodAnalysisResult(template_conflicts, conflicts, classifications, sorted(rules_by_key.values(), key=lambda row: (row["Sistema"], row["ID da regra"])), embedded, gaps, user_impacted, no_direct, sat_candidates, summary)

