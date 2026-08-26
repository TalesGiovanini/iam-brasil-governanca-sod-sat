from __future__ import annotations

import math
import warnings
import zipfile
from xml.etree import ElementTree
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import column_index_from_string


def _normalise(value):
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return None
    if hasattr(value, "item"):
        try:
            value = value.item()
        except ValueError:
            pass
    return value


def _write_dataset(sheet, spec: dict, records: list[dict]) -> None:
    """Preenche a Ã¡rea de dados em uma cÃ³pia do modelo oficial.

    A gravaÃ§Ã£o Ã© delegada ao openpyxl, em vez de alterar XML diretamente. Isso
    evita que o Excel repare o arquivo e descarte abas quando o modelo contÃ©m
    recursos visuais do Office.
    """
    start = int(spec["linha_inicial_dados"])
    header = int(spec["linha_cabecalho"])
    columns = spec["colunas"]
    style_source = {
        column: copy(sheet.cell(start, column_index_from_string(column))._style)
        for column in columns.values()
    }
    row_height = sheet.row_dimensions[start].height
    if sheet.max_row >= start:
        sheet.delete_rows(start, sheet.max_row - start + 1)
    for offset, record in enumerate(records):
        row = start + offset
        for field, column in columns.items():
            cell = sheet.cell(row, column_index_from_string(column), _normalise(record.get(field)))
            cell._style = copy(style_source[column])
        if row_height:
            sheet.row_dimensions[row].height = row_height
    last_column = max(columns.values(), key=column_index_from_string)
    last_row = max(header, start + len(records) - 1)
    sheet.auto_filter.ref = f"A{header}:{last_column}{last_row}"


def _system_label(datasets: dict[str, list[dict]], supplied_label: str | None = None) -> str:
    """ObtÃ©m um rÃ³tulo editorial sem alterar o sistema informado na fonte."""
    if supplied_label and supplied_label.strip():
        return supplied_label.strip()
    systems = {
        str(record.get("sistema")).strip()
        for records in datasets.values()
        for record in records
        if record.get("sistema") is not None and str(record.get("sistema")).strip() not in {"", "-"}
    }
    if len(systems) == 1:
        return next(iter(systems))
    if len(systems) > 1:
        return "MÃºltiplos sistemas"
    return "Sistema nÃ£o identificado"


def _remove_header_merges(sheet, final_column: int) -> None:
    """Libera somente a Ã¡rea da faixa institucional que serÃ¡ recomposta."""
    for merged in list(sheet.merged_cells.ranges):
        if merged.min_row <= 3 and merged.max_row <= 3 and merged.min_col <= final_column:
            sheet.unmerge_cells(str(merged))


def _apply_brand_header(sheet, label: str, logo_path: Path) -> None:
    """RecompÃµe o cabeÃ§alho que o openpyxl nÃ£o consegue preservar em DrawingML.

    O modelo de origem contÃ©m formas do Office. Essas formas sÃ£o descartadas pelo
    leitor do Python e eram a causa de uma planilha visualmente incompleta. A
    faixa abaixo usa cÃ©lulas e uma imagem PNG, recursos nativos e compatÃ­veis com
    Excel, sem depender de vÃ­nculos externos.
    """
    final_column = max(sheet.max_column, 6)
    _remove_header_merges(sheet, final_column)
    blue, cyan = "3953C7", "49BBD7"
    for row in range(1, 4):
        for column in range(1, final_column + 1):
            cell = sheet.cell(row, column)
            cell.fill = PatternFill("solid", fgColor=blue if row < 3 else cyan)
            cell.value = None
    sheet.merge_cells(start_row=1, start_column=2, end_row=2, end_column=final_column)
    sheet.merge_cells(start_row=3, start_column=2, end_row=3, end_column=final_column)
    title = sheet.cell(1, 2, "IAM Brasil | GovernanÃ§a de Acessos")
    title.font = Font(name="Aptos Display", size=18, bold=True, color="FFFFFF")
    title.alignment = Alignment(vertical="center")
    subtitle = sheet.cell(3, 2, f"Matriz SoD e SAT â€” {label}")
    subtitle.font = Font(name="Aptos", size=11, bold=True, color="FFFFFF")
    subtitle.alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 25
    sheet.row_dimensions[2].height = 25
    sheet.row_dimensions[3].height = 20
    if logo_path.is_file():
        logo = ExcelImage(str(logo_path))
        logo.width, logo.height = 52, 52
        sheet.add_image(logo, "A1")


def _apply_sheet_identity(workbook, template: Path, template_config: dict, datasets: dict[str, list[dict]], system_label: str | None) -> None:
    """Padroniza a identidade visual e o sistema analisado em todas as abas."""
    label = _system_label(datasets, system_label)
    logo_path = template.parent.parent / "recursos" / "iam_brasil_access_governance.png"
    for dataset, spec in template_config["abas_saida"].items():
        sheet = workbook[spec["aba"]]
        # A aba de conflitos inicia na linha 1 com sua prÃ³pria tabela. NÃ£o hÃ¡
        # espaÃ§o seguro para uma faixa sem deslocar ou apagar os dados.
        if dataset != "atividades_conflitantes":
            _apply_brand_header(sheet, label, logo_path)
        # As abas Matriz e SAT jÃ¡ reservam a linha 4 para o tÃ­tulo editorial.
        if dataset in {"matriz_funcional", "sat"}:
            final_column = max(sheet.max_column, column_index_from_string(max(spec["colunas"].values(), key=column_index_from_string)))
            for merged in list(sheet.merged_cells.ranges):
                if merged.min_row == 4 and merged.max_row == 4:
                    sheet.unmerge_cells(str(merged))
            sheet.merge_cells(start_row=4, start_column=1, end_row=4, end_column=final_column)
            title = "Matriz Funcional" if dataset == "matriz_funcional" else "SAT â€” TransaÃ§Ãµes de Acesso SensÃ­vel"
            cell = sheet.cell(4, 1, f"{title} | Sistema analisado: {label}")
            cell.font = Font(name="Aptos", size=14, bold=True, color="00A6D6")
            cell.alignment = Alignment(vertical="center")
            sheet.row_dimensions[4].height = 26
        sheet.sheet_properties.tabColor = "3953C7"


def _validate_exported_data(output: Path, template_config: dict, datasets: dict[str, list[dict]]) -> None:
    """Valida ZIP e conteÃºdo persistido antes de a interface informar sucesso."""
    with zipfile.ZipFile(output) as archive:
        broken_part = archive.testzip()
        if broken_part:
            raise ValueError(f"Arquivo XLSX invÃ¡lido apÃ³s a exportaÃ§Ã£o: {broken_part}")
        external_parts = [name for name in archive.namelist() if name.startswith("xl/externalLinks/")]
        if external_parts:
            raise ValueError("Arquivo XLSX contÃ©m vÃ­nculos externos que acionariam reparo no Excel.")
        rels = archive.read("xl/_rels/workbook.xml.rels")
        if b"externalLink" in rels:
            raise ValueError("Arquivo XLSX contÃ©m relacionamento externo que acionaria reparo no Excel.")
        root = ElementTree.fromstring(archive.read("xl/workbook.xml"))
        names = root.findall("{http://schemas.openxmlformats.org/spreadsheetml/2006/main}definedNames/{http://schemas.openxmlformats.org/spreadsheetml/2006/main}definedName")
        unsafe_names = [item.get("name", "") for item in names if item.get("name") != "_xlnm._FilterDatabase"]
        if unsafe_names:
            raise ValueError("Arquivo XLSX contÃ©m intervalos nomeados herdados do template que poderiam acionar reparo no Excel.")
    workbook = load_workbook(output, read_only=True, data_only=False)
    try:
        for dataset, spec in template_config["abas_saida"].items():
            records = datasets.get(dataset, [])
            if not records:
                continue
            sheet = workbook[spec["aba"]]
            start = int(spec["linha_inicial_dados"])
            expected_last = start + len(records) - 1
            if sheet.max_row < expected_last:
                raise ValueError(f"ValidaÃ§Ã£o da aba '{sheet.title}' falhou: esperadas {len(records)} linhas de dados.")
            values = next(sheet.iter_rows(min_row=start, max_row=start, values_only=True))
            if not any(value is not None and str(value) != "" for value in values):
                raise ValueError(f"ValidaÃ§Ã£o da aba '{sheet.title}' falhou: a primeira linha gravada estÃ¡ vazia.")
    finally:
        workbook.close()


def export_template(template: Path, output: Path, template_config: dict, datasets: dict[str, list[dict]], system_label: str | None = None) -> None:
    """Cria uma cÃ³pia compatÃ­vel com Excel a partir do template oficial preservado."""
    if template.resolve() == output.resolve():
        raise ValueError("O arquivo de saÃ­da nÃ£o pode sobrescrever o template oficial.")
    output.parent.mkdir(parents=True, exist_ok=True)
    # keep_links=False impede que a cÃ³pia carregue os /xl/externalLinks do
    # modelo. Esses vÃ­nculos eram o conteÃºdo que o Excel reparava na abertura.
    with warnings.catch_warnings():
        warnings.filterwarnings("ignore", message="DrawingML support is incomplete.*")
        workbook = load_workbook(template, keep_links=False)
    try:
        workbook._external_links = []
        # O modelo recebido tem centenas de nomes definidos de outros arquivos
        # e planilhas jÃ¡ inexistentes. Eles nÃ£o fazem parte da matriz SoD/SAT e
        # o Excel os remove ao abrir, emitindo o aviso de reparo. A saÃ­da Ã© uma
        # nova cÃ³pia operacional, portanto nÃ£o deve herdar esse lixo tÃ©cnico.
        workbook.defined_names.clear()
        for sheet in workbook.worksheets:
            sheet.defined_names.clear()
        for dataset, spec in template_config["abas_saida"].items():
            sheet_name = spec["aba"]
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"A aba oficial '{sheet_name}' nÃ£o foi localizada no template.")
            _write_dataset(workbook[sheet_name], spec, datasets.get(dataset, []))
        _apply_sheet_identity(workbook, template, template_config, datasets, system_label)
        workbook.save(output)
    finally:
        workbook.close()
    _validate_exported_data(output, template_config, datasets)


